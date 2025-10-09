from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
import pandas as pd

df = pd.read_csv("sales.csv")
print(df)

df = df.drop(columns=["extra_column"])

df.rename(columns={"units_sold":"Units Sold", "unit_price":"Unit Price", "product":"Product"}, inplace=True)

df = df[df["Units Sold"] > 5]

df["Revenue"] = df["Units Sold"] * df["Unit Price"]

print(df)

# Convert DataFrame to Excel
wb = Workbook()
ws = wb.active
ws.title = "Sales Report"

# Add header row
header = list(df.columns)
ws.append(header)

# Style header
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data rows
for row in df.itertuples(index=False):
    ws.append(row)

# Format currency columns
for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):  # Unit Price + Revenue
    for cell in row:
        cell.number_format = '"$"#,##0.00'
        cell.alignment = Alignment(horizontal="center", vertical="center")

for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    if i % 2 == 0:
        for cell in row:
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


# Save file
wb.save("formatted_sales_report.xlsx")
