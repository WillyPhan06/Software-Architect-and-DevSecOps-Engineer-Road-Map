from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import random

xlsx_file = r"D:\GitHub_Repos\Personal_GitHub_Repos\Software-Architect-and-DevSecOps-Engineer-Road-Map\Topics\Python_Automation\Mini_Project\day_9_mp\employee_data.xlsx"

# Load the Excel file
wb = load_workbook(xlsx_file)
ws = wb.active

# Get all salaries, ages, and departments
salaries = [cell.value for cell in ws['D'][1:]]  # skip header
ages = [cell.value for cell in ws['B'][1:]]      # skip header
departments = list(set(cell.value for cell in ws['C'][1:]))  # unique departments
countries = list(set(cell.value for cell in ws['E'][1:]))    # unique countries

# Calculate average salary
avg_salary = 7500

# Define department colors
dept_colors = {}
color_palette = ["FFB6C1", "ADD8E6", "90EE90", "FFA500", "DDA0DD"]  # arbitrary colors
for i, dept in enumerate(departments):
    dept_colors[dept] = color_palette[i % len(color_palette)]

# Define country colors
country_colors = {}
country_palette = ["FFFF99", "99FF99", "FF9999", "99CCFF", "FFCC99"]
for i, country in enumerate(countries):
    country_colors[country] = country_palette[i % len(country_palette)]

# Salary shades
red_shades = ["FFC7CE", "FF9999", "FF6666", "FF0000"]   # increasing red
green_shades = ["C6EFCE", "99FF99", "33CC33", "008000"] # increasing green

# Get min and max ages
min_age = min(ages)
max_age = max(ages)

# Start applying formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    salary = row[3].value
    age = row[1].value
    department = row[2].value
    country = row[4].value
    sex = row[5].value

    # ---- Salary coloring ----
    if salary < avg_salary:
        # pick shade based on how low it is
        if salary < avg_salary*0.5:
            fill = PatternFill(start_color=red_shades[3], end_color=red_shades[3], fill_type="solid")
        elif salary < avg_salary*0.65:
            fill = PatternFill(start_color=red_shades[2], end_color=red_shades[2], fill_type="solid")
        elif salary < avg_salary*0.85:
            fill = PatternFill(start_color=red_shades[1], end_color=red_shades[1], fill_type="solid")
        else:
            fill = PatternFill(start_color=red_shades[0], end_color=red_shades[0], fill_type="solid")
    else:
        # pick shade based on how high it is
        if salary > avg_salary*1.5:
            fill = PatternFill(start_color=green_shades[3], end_color=green_shades[3], fill_type="solid")
        elif salary > avg_salary*1.25:
            fill = PatternFill(start_color=green_shades[2], end_color=green_shades[2], fill_type="solid")
        elif salary > avg_salary*1.05:
            fill = PatternFill(start_color=green_shades[1], end_color=green_shades[1], fill_type="solid")
        else:
            fill = PatternFill(start_color=green_shades[0], end_color=green_shades[0], fill_type="solid")
    row[3].fill = fill

    # ---- Age coloring ----
    if age == min_age:
        row[1].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # green
    elif age == max_age:
        row[1].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # red
    else:
        row[1].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # yellow

    # ---- Department coloring ----
    row[2].fill = PatternFill(start_color=dept_colors[department], end_color=dept_colors[department], fill_type="solid")

    # ---- Country coloring ----
    row[4].fill = PatternFill(start_color=country_colors[country], end_color=country_colors[country], fill_type="solid")

    # ---- Sex coloring ----
    if sex.upper() == "F":
        row[5].fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # pink
    else:
        row[5].fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # blue

# Save the formatted file
wb.save(xlsx_file)
print("Generated with conditional coloring!")
