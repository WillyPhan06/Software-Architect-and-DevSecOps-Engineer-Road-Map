from openpyxl import Workbook, load_workbook
from dataclasses import dataclass

xlsx_path = r"D:\GitHub_Repos\Personal_GitHub_Repos\Software-Architect-and-DevSecOps-Engineer-Road-Map\Topics\Python_Automation\Mini_Project\day_8_mp\products.xlsx"

wb = load_workbook(xlsx_path)
ws = wb.active

@dataclass
class Product:
    product_id: int
    name: str
    category: str
    price: float
    stock: int

products = []

for row in list(ws.iter_rows(values_only=True))[1:]:
    product_id, name, category, price, stock = row
    products.append(Product(product_id, name, category, price, stock))

for product in products:
    print("----------")
    print(product.product_id)
    print(product.name)
    print(product.category)
    print(product.price)
    print(product.stock)



