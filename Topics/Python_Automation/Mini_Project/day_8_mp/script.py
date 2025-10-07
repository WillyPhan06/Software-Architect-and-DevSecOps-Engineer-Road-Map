from openpyxl import Workbook, load_workbook
import os
import random

xlsx_path = r"D:\GitHub_Repos\Personal_GitHub_Repos\Software-Architect-and-DevSecOps-Engineer-Road-Map\Topics\Python_Automation\Mini_Project\day_8_mp\products.xlsx"

# Sample data
product_names = ["Keyboard", "Mouse", "Monitor", "Laptop", "Headphones", "Microphone", "Webcam", "Charger", "USB Cable", "Desk Lamp"]
categories = ["Electronics", "Accessories", "Peripherals", "Office", "Gadgets"]

# Check if the file exists
if os.path.exists(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
ws.append(["Product ID", "Name", "Category", "Price", "Stock"])  # header

# Generate random rows
for i in range(5):  # Add 5 random products
    product_id = ws.max_row  # simple way to make incremental IDs
    name = random.choice(product_names)
    category = random.choice(categories)
    price = round(random.uniform(10, 500), 2)
    stock = random.randint(1, 200)

    ws.append([product_id, name, category, price, stock])

# Save workbook
wb.save(xlsx_path)
print(f"5 random products added to {xlsx_path}")
