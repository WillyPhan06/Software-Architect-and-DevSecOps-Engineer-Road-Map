from openpyxl import load_workbook

xlsx_path_new = r"D:\GitHub_Repos\Personal_GitHub_Repos\Software-Architect-and-DevSecOps-Engineer-Road-Map\Topics\Python_Automation\Code_Practice\day_8\staff.xlsx"


wb = load_workbook(xlsx_path_new)
ws = wb["Staff"]

# Read all rows
for row in ws.iter_rows(values_only=True):
    print(row)

# Return each row as an actual cell object and access its attribute
for row in ws.iter_rows(values_only=False):
    print("\n")
    for cell in row:
        print(cell.coordinate, end=", ")  # e.g. 'A1'
        print(cell.value, end=", ")       # actual value 
        # print(cell.font.bold, end=". ") # styling info

print("\n")
print("-------------------------------------------------------------")


count = 0  # to count salary > 500
sigmas = []

# Skip the header row by slicing [1:]
for row in list(ws.iter_rows(values_only=True))[1:]:
    name, age, department, salary = row
    print(name)

    if salary > 500:
        sigmas.append(name)
        count += 1

print("\nNumber of rows with salary > 500:", count)
print("\nHere are the sigmas: ")
for sigma in sigmas:
    print(sigma)

print("\n")
print("-------------------------------------------------------------")

ws["A2"] = "Alicu"     # Change Alice’s name
ws["B3"] = 69     # Change Bob's age
wb.save(xlsx_path_new)
print("Changed stuff")
