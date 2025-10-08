from openpyxl import load_workbook
from openpyxl.styles import Font

upcoming_path = r"D:\GitHub_Repos\Personal_GitHub_Repos\Software-Architect-and-DevSecOps-Engineer-Road-Map\Topics\Python_Automation\Code_Practice\day_9\employees.xlsx"
after_style_path = r"D:\GitHub_Repos\Personal_GitHub_Repos\Software-Architect-and-DevSecOps-Engineer-Road-Map\Topics\Python_Automation\Code_Practice\day_9\after_style.xlsx"


wb = load_workbook(upcoming_path)
ws = wb.active

# Bold header row
for cell in ws[1]:
    cell.font = Font(bold=True, color="00FF00", size=12, name="Calibri")

for cell in ws['A'][1:]:
    cell.font = Font(bold=True, color="FF0000", size=12, name="Calibri")

for col in ws[1]:
    if col.value == "Age":
        age_column_letter = col.column_letter

for cell in ws[age_column_letter][1:]:
    cell.font = Font(italic=True, size=12, name="Calibri")

wb.save(after_style_path)

print("Finished styling")
