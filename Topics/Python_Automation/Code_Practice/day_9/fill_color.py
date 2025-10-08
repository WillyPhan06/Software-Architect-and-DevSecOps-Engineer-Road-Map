from openpyxl.styles import PatternFill, Alignment, GradientFill
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import numbers

upcoming_path = r"D:\GitHub_Repos\Personal_GitHub_Repos\Software-Architect-and-DevSecOps-Engineer-Road-Map\Topics\Python_Automation\Code_Practice\day_9\employees.xlsx"
after_style_path = r"D:\GitHub_Repos\Personal_GitHub_Repos\Software-Architect-and-DevSecOps-Engineer-Road-Map\Topics\Python_Automation\Code_Practice\day_9\after_style.xlsx"


wb = load_workbook(upcoming_path)
ws = wb.active


# Fill header background
for cell in ws[1]:
    cell.fill = GradientFill(type="linear", stop=("4F81BD", "00B050"))
    cell.alignment = Alignment(horizontal="center", vertical="center")

for cell in ws['C']:
    if cell.value.strip() == "IT":
        cell.font = Font(color="003366")
    elif cell.value.strip() == "HR":
        cell.font = Font(color="2ECC71")
    elif cell.value.strip() == "Marketing":
        cell.font = Font(color="00AEEF")

for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.number_format = '"$"#,##0.00'



wb.save(after_style_path)

print("Finished adding blue shade or gradient color")
